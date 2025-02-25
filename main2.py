import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


class ExtratorBeneficiariosApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Extrator de Beneficiários - Engemon")
        self.root.geometry("600x400")
        self.root.configure(bg="#f0f0f0")
        
        # Título da aplicação
        titulo = tk.Label(root, text="Extrator de Beneficiários", font=("Arial", 18, "bold"), bg="#f0f0f0")
        titulo.pack(pady=20)
        
        # Descrição
        descricao = tk.Label(root, text="Selecione o arquivo PDF contendo a tabela de beneficiários", font=("Arial", 12), bg="#f0f0f0")
        descricao.pack(pady=10)
        
        # Frame para o botão de upload
        frame_botao = tk.Frame(root, bg="#f0f0f0")
        frame_botao.pack(pady=20)
        
        # Botão para selecionar arquivo
        self.btn_selecionar = tk.Button(frame_botao, text="Selecionar PDF", command=self.selecionar_arquivo, 
                                    font=("Arial", 12), bg="#4CAF50", fg="white", width=20, height=2)
        self.btn_selecionar.pack(side=tk.LEFT, padx=10)
        
        # Botão de processamento
        self.btn_processar = tk.Button(frame_botao, text="Processar", command=self.processar_arquivo, 
                                   font=("Arial", 12), bg="#2196F3", fg="white", width=20, height=2)
        self.btn_processar.pack(side=tk.LEFT, padx=10)
        self.btn_processar.config(state=tk.DISABLED)
        
        # Label para mostrar o arquivo selecionado
        self.label_arquivo = tk.Label(root, text="Nenhum arquivo selecionado", font=("Arial", 10), bg="#f0f0f0")
        self.label_arquivo.pack(pady=10)
        
        # Status bar
        self.status_bar = tk.Label(root, text="Aguardando...", font=("Arial", 10), bg="#e0e0e0", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Variável para armazenar o caminho do arquivo selecionado
        self.arquivo_pdf = None

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(title="Selecione o arquivo PDF", 
                                            filetypes=[("Arquivos PDF", "*.pdf")])
        if arquivo:
            self.arquivo_pdf = arquivo
            nome_arquivo = os.path.basename(arquivo)
            self.label_arquivo.config(text=f"Arquivo selecionado: {nome_arquivo}")
            self.btn_processar.config(state=tk.NORMAL)
            self.status_bar.config(text="Arquivo selecionado. Clique em Processar para continuar.")

    def processar_arquivo(self):
        if not self.arquivo_pdf:
            messagebox.showerror("Erro", "Nenhum arquivo selecionado!")
            return
        
        self.status_bar.config(text="Processando o arquivo PDF...")
        self.root.update()
        
        try:
            # Extrair dados do PDF
            dados = self.extrair_dados_pdf(self.arquivo_pdf)
            
            if not dados:
                messagebox.showerror("Erro", "Não foi possível extrair dados da tabela no PDF.")
                self.status_bar.config(text="Erro ao processar o arquivo.")
                return
            
            # Criar nome do arquivo Excel
            nome_base = os.path.splitext(os.path.basename(self.arquivo_pdf))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_excel = f"{nome_base}_extraido_{timestamp}.xlsx"
            caminho_excel = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivos Excel", "*.xlsx")],
                initialfile=nome_excel
            )
            
            if not caminho_excel:
                self.status_bar.config(text="Operação cancelada.")
                return
            
            # Processar dados e adicionar campos calculados
            df = self.processar_dados(dados)
            
            # Salvar para Excel com formatação
            self.salvar_excel_formatado(df, caminho_excel)
            
            messagebox.showinfo("Sucesso", f"Dados extraídos com sucesso!\nArquivo salvo em: {caminho_excel}")
            self.status_bar.config(text=f"Concluído. Arquivo Excel salvo.")
            
            # Perguntar se deseja abrir o arquivo Excel
            if messagebox.askyesno("Abrir arquivo", "Deseja abrir o arquivo Excel gerado?"):
                os.startfile(caminho_excel)
        
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processamento: {str(e)}")
            self.status_bar.config(text="Erro durante o processamento.")

    def extrair_dados_pdf(self, caminho_pdf):
        dados_extraidos = []
        
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                tabelas = pagina.extract_tables()
                
                for tabela in tabelas:
                    # Verifica se a tabela tem cabeçalho
                    if not tabela or len(tabela) <= 1:
                        continue
                    
                    # Identifica o cabeçalho
                    cabecalho = tabela[0]
                    
                    # Verifica se é a tabela correta (procurando por colunas específicas)
                    colunas_esperadas = ['Status', 'Tipo', 'Nome', 'CPF', 'Nascimento', 'Titular', 'Parentesco', 'Matricula', 'Data de inclusão']
                    if not all(col in ' '.join(cabecalho) for col in ['Status', 'CPF', 'Matricula']):
                        continue
                    
                    # Mapeia índices das colunas necessárias
                    indices_colunas = {}
                    for i, col in enumerate(cabecalho):
                        for col_esperada in colunas_esperadas:
                            if col and col_esperada.lower() in col.lower():
                                indices_colunas[col_esperada] = i
                                break
                    
                    # Processa cada linha da tabela (exceto o cabeçalho)
                    for linha in tabela[1:]:
                        if all(cell is None or cell.strip() == '' for cell in linha):
                            continue  # Pula linhas vazias
                        
                        registro = {}
                        for col_nome, col_idx in indices_colunas.items():
                            if col_idx < len(linha):
                                valor = linha[col_idx]
                                registro[col_nome] = valor.strip() if valor else ''
                        
                        # Pular se os dados principais estiverem vazios
                        if not registro.get('Nome') and not registro.get('CPF'):
                            continue
                            
                        dados_extraidos.append(registro)
        
        return dados_extraidos

    def processar_dados(self, dados):
        # Converter lista de dicionários para DataFrame
        df = pd.DataFrame(dados)
        
        # Converter e normalizar datas de nascimento
        df['Data Nascimento Original'] = df['Nascimento']  # Guardar valor original
        
        # Tentar converter as datas usando vários formatos comuns
        formatos_data = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y']
        
        def converter_data(data_str):
            if not data_str or pd.isna(data_str):
                return None
                
            data_str = str(data_str).strip()
            
            # Tentar cada formato
            for formato in formatos_data:
                try:
                    return pd.to_datetime(data_str, format=formato)
                except:
                    continue
            
            # Se nenhum formato funcionou, tentar método genérico com dayfirst=True
            try:
                return pd.to_datetime(data_str, dayfirst=True)
            except:
                return None
        
        df['Nascimento_dt'] = df['Nascimento'].apply(converter_data)
        
        # Calcular idade - método robusto
        def calcular_idade(data_nasc):
            if pd.isna(data_nasc):
                return None
                
            hoje = datetime.now()
            return hoje.year - data_nasc.year - ((hoje.month, hoje.day) < (data_nasc.month, data_nasc.day))
        
        df['Idade'] = df['Nascimento_dt'].apply(calcular_idade)
        
        # Adicionar coluna Status Idade
        df['Status Idade'] = df['Idade'].apply(
            lambda x: 'AVISO' if pd.notna(x) and x > 22 else 'OK' if pd.notna(x) else ''
        )
        
        # Formatar CPF (adicionar máscaras)
        df['CPF'] = df['CPF'].apply(self.formatar_cpf)
        
        # Tratar data de inclusão
        df['Data de inclusão'] = df['Data de inclusão'].apply(converter_data)
        
        # Formatar datas para exibição
        df['Nascimento'] = df['Nascimento_dt'].apply(
            lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else df.loc[df.index[df['Nascimento_dt'] == x].tolist()[0], 'Data Nascimento Original'] if len(df.index[df['Nascimento_dt'] == x].tolist()) > 0 else ''
        )
        
        df['Data de inclusão'] = df['Data de inclusão'].apply(
            lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else ''
        )
        
        # Remover colunas temporárias
        df = df.drop(['Nascimento_dt', 'Data Nascimento Original'], axis=1, errors='ignore')
        
        # Reorganizar colunas
        colunas_ordem = [col for col in df.columns if col not in ['Idade', 'Status Idade']] + ['Idade', 'Status Idade']
        df = df[colunas_ordem]
        
        return df
    
    def formatar_cpf(self, cpf):
        """Formata um CPF com a máscara xxx.xxx.xxx-xx"""
        if not cpf or pd.isna(cpf):
            return ''
            
        # Remover caracteres não numéricos
        cpf_numeros = re.sub(r'\D', '', str(cpf))
        
        # Verificar se tem 11 dígitos
        if len(cpf_numeros) != 11:
            return cpf  # Retorna original se não tiver 11 dígitos
            
        # Aplicar máscara
        return f"{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}"

    def salvar_excel_formatado(self, df, caminho_excel):
        # Substituir valores NaN por string vazia
        df = df.fillna('')
        
        # Salvar DataFrame em Excel
        with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Beneficiários')
        
        # Carregar o arquivo Excel para formatação
        wb = load_workbook(caminho_excel)
        ws = wb['Beneficiários']
        
        # Definir estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Formatar linhas de dados
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Alinhar colunas numéricas à direita
                if col in [ws.cell(row=1, column=col).column_letter for col in range(1, max_col + 1) 
                          if ws.cell(row=1, column=col).value in ['CPF', 'Matricula', 'Idade']]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                # Alinhar datas ao centro
                if col in [ws.cell(row=1, column=col).column_letter for col in range(1, max_col + 1) 
                          if ws.cell(row=1, column=col).value in ['Nascimento', 'Data de inclusão']]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatar Status Idade com cores
                if ws.cell(row=1, column=col).value == 'Status Idade':
                    if cell.value == 'OK':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(name='Arial', size=11, color="006100", bold=True)
                    elif cell.value == 'AVISO':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(name='Arial', size=11, color="9C0006", bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar largura das colunas automaticamente
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        # Congelar o painel no cabeçalho
        ws.freeze_panes = "A2"
        
        # Adicionar filtro no cabeçalho
        ws.auto_filter.ref = ws.dimensions
        
        # Salvar as alterações
        wb.save(caminho_excel)

def main():
    root = tk.Tk()
    app = ExtratorBeneficiariosApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()